"""
Module 8: Output Generator (output_generator.py)
Purpose: Generate required output files (JSON, Excel, verification report)
"""

import json
import pandas as pd
from datetime import datetime
from pathlib import Path
import logging
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class OutputGenerator:
    """Handles generation of all output files in required formats."""
    
    def __init__(self, output_dir: str = "output"):
        """
        Initialize OutputGenerator.
        
        Args:
            output_dir: Directory path for output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        # Create subdirectories
        self.seal_dir = self.output_dir / "seal_signatures"
        self.seal_dir.mkdir(exist_ok=True)
    
    def format_output_structure(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Structure extracted data according to output requirements.
        
        Args:
            raw_data: Raw extracted data from all modules
            
        Returns:
            Formatted data structure for output
        """
        try:
            # Basic invoice information
            invoice_info = {
                "invoice_number": raw_data.get("invoice_number", ""),
                "invoice_date": raw_data.get("invoice_date", ""),
                "po_number": raw_data.get("po_number", ""),
                "vendor_name": raw_data.get("vendor_name", ""),
                "vendor_gst": raw_data.get("vendor_gst", ""),
                "buyer_name": raw_data.get("buyer_name", ""),
                "buyer_gst": raw_data.get("buyer_gst", ""),
                "shipping_address": raw_data.get("shipping_address", {})
            }
            
            # Line items formatting
            line_items = []
            raw_line_items = raw_data.get("line_items", [])
            
            for item in raw_line_items:
                formatted_item = {
                    "description": item.get("description", ""),
                    "hsn_code": item.get("hsn_code", ""),
                    "quantity": self._safe_float(item.get("quantity", 0)),
                    "unit_price": self._safe_float(item.get("unit_price", 0)),
                    "total_amount": self._safe_float(item.get("total_amount", 0)),
                    "gst_rate": self._safe_float(item.get("gst_rate", 0)),
                    "gst_amount": self._safe_float(item.get("gst_amount", 0))
                }
                line_items.append(formatted_item)
            
            # Financial totals
            totals = {
                "subtotal": self._safe_float(raw_data.get("subtotal", 0)),
                "discount": self._safe_float(raw_data.get("discount", 0)),
                "total_gst": self._safe_float(raw_data.get("total_gst", 0)),
                "final_total": self._safe_float(raw_data.get("final_total", 0))
            }
            
            # Seal/signature information
            seals_signatures = {
                "detected": raw_data.get("seals_detected", False),
                "count": raw_data.get("seal_count", 0),
                "file_paths": raw_data.get("seal_file_paths", [])
            }
            
            # Metadata
            metadata = {
                "extraction_timestamp": datetime.now().isoformat(),
                "source_file": raw_data.get("source_file", ""),
                "total_pages": raw_data.get("total_pages", 1),
                "processing_time": raw_data.get("processing_time", 0)
            }
            
            formatted_data = {
                "invoice_info": invoice_info,
                "line_items": line_items,
                "totals": totals,
                "seals_signatures": seals_signatures,
                "metadata": metadata
            }
            
            logger.info("Data structure formatted successfully")
            return formatted_data
            
        except Exception as e:
            logger.error(f"Error formatting output structure: {str(e)}")
            raise
    
    def create_json_output(self, extracted_data: Dict[str, Any], 
                          output_path: Optional[str] = None) -> str:
        """
        Generate extracted_data.json file.
        
        Args:
            extracted_data: Formatted extracted data
            output_path: Custom output path (optional)
            
        Returns:
            Path to generated JSON file
        """
        try:
            if output_path is None:
                output_path = self.output_dir / "extracted_data.json"
            else:
                output_path = Path(output_path)
            
            # Format the data structure
            formatted_data = self.format_output_structure(extracted_data)
            
            # Write JSON file with proper formatting
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(formatted_data, f, indent=2, ensure_ascii=False, 
                         default=self._json_serializer)
            
            logger.info(f"JSON output created: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error creating JSON output: {str(e)}")
            raise
    
    def create_excel_output(self, extracted_data: Dict[str, Any], 
                           output_path: Optional[str] = None) -> str:
        """
        Generate extracted_data.xlsx file with multiple sheets.
        
        Args:
            extracted_data: Formatted extracted data
            output_path: Custom output path (optional)
            
        Returns:
            Path to generated Excel file
        """
        try:
            if output_path is None:
                output_path = self.output_dir / "extracted_data.xlsx"
            else:
                output_path = Path(output_path)
            
            # Format the data structure
            formatted_data = self.format_output_structure(extracted_data)
            
            # Create Excel workbook
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Invoice Summary
                self._create_invoice_summary_sheet(formatted_data, writer)
                
                # Sheet 2: Line Items
                self._create_line_items_sheet(formatted_data, writer)
                
                # Sheet 3: Verification Details
                if 'verification_data' in extracted_data:
                    self._create_verification_sheet(
                        extracted_data['verification_data'], writer
                    )
            
            # Apply formatting
            self._format_excel_file(output_path)
            
            logger.info(f"Excel output created: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error creating Excel output: {str(e)}")
            raise
    
    def create_verification_report(self, verification_data: Dict[str, Any], 
                                 output_path: Optional[str] = None) -> str:
        """
        Generate verifiability_report.json file.
        
        Args:
            verification_data: Verification results and confidence scores
            output_path: Custom output path (optional)
            
        Returns:
            Path to generated verification report
        """
        try:
            if output_path is None:
                output_path = self.output_dir / "verifiability_report.json"
            else:
                output_path = Path(output_path)
            
            # Structure verification report
            report = {
                "overall_confidence": verification_data.get("overall_confidence", 0.0),
                "verification_status": verification_data.get("verification_status", "PENDING"),
                "timestamp": datetime.now().isoformat(),
                
                "field_confidence_scores": {
                    "invoice_number": verification_data.get("invoice_number_confidence", 0.0),
                    "invoice_date": verification_data.get("invoice_date_confidence", 0.0),
                    "vendor_info": verification_data.get("vendor_info_confidence", 0.0),
                    "line_items": verification_data.get("line_items_confidence", 0.0),
                    "totals": verification_data.get("totals_confidence", 0.0)
                },
                
                "mathematical_verification": {
                    "line_item_calculations": verification_data.get("line_item_calc_valid", True),
                    "subtotal_verification": verification_data.get("subtotal_valid", True),
                    "total_verification": verification_data.get("total_valid", True),
                    "gst_calculations": verification_data.get("gst_calc_valid", True)
                },
                
                "data_quality_flags": {
                    "missing_fields": verification_data.get("missing_fields", []),
                    "low_confidence_fields": verification_data.get("low_confidence_fields", []),
                    "calculation_errors": verification_data.get("calculation_errors", []),
                    "format_issues": verification_data.get("format_issues", [])
                },
                
                "seal_signature_verification": {
                    "seals_detected": verification_data.get("seals_detected", False),
                    "signature_count": verification_data.get("signature_count", 0),
                    "authenticity_score": verification_data.get("authenticity_score", 0.0)
                },
                
                "recommendations": verification_data.get("recommendations", []),
                
                "processing_metadata": {
                    "ocr_engine": verification_data.get("ocr_engine", ""),
                    "processing_time": verification_data.get("processing_time", 0),
                    "image_quality_score": verification_data.get("image_quality_score", 0.0)
                }
            }
            
            # Write verification report
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, ensure_ascii=False, 
                         default=self._json_serializer)
            
            logger.info(f"Verification report created: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error creating verification report: {str(e)}")
            raise
    
    def generate_all_outputs(self, extracted_data: Dict[str, Any], 
                           verification_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Generate all required output files.
        
        Args:
            extracted_data: Extracted invoice data
            verification_data: Verification results
            
        Returns:
            Dictionary with paths to all generated files
        """
        try:
            output_paths = {}
            
            # Generate JSON output
            output_paths['json'] = self.create_json_output(extracted_data)
            
            # Generate Excel output
            output_paths['excel'] = self.create_excel_output(extracted_data)
            
            # Generate verification report
            output_paths['verification'] = self.create_verification_report(verification_data)
            
            # Copy seal/signature files if they exist
            if extracted_data.get('seal_file_paths'):
                output_paths['seals'] = str(self.seal_dir)
            
            logger.info("All output files generated successfully")
            return output_paths
            
        except Exception as e:
            logger.error(f"Error generating outputs: {str(e)}")
            raise
    
    def _create_invoice_summary_sheet(self, data: Dict[str, Any], writer):
        """Create invoice summary sheet in Excel."""
        invoice_info = data['invoice_info']
        totals = data['totals']
        metadata = data['metadata']
        
        summary_data = [
            ['Field', 'Value'],
            ['Invoice Number', invoice_info.get('invoice_number', '')],
            ['Invoice Date', invoice_info.get('invoice_date', '')],
            ['PO Number', invoice_info.get('po_number', '')],
            ['Vendor Name', invoice_info.get('vendor_name', '')],
            ['Vendor GST', invoice_info.get('vendor_gst', '')],
            ['Buyer Name', invoice_info.get('buyer_name', '')],
            ['Buyer GST', invoice_info.get('buyer_gst', '')],
            ['', ''],  # Empty row
            ['Financial Summary', ''],
            ['Subtotal', totals.get('subtotal', 0)],
            ['Discount', totals.get('discount', 0)],
            ['Total GST', totals.get('total_gst', 0)],
            ['Final Total', totals.get('final_total', 0)],
            ['', ''],  # Empty row
            ['Processing Info', ''],
            ['Extraction Date', metadata.get('extraction_timestamp', '')],
            ['Source File', metadata.get('source_file', '')],
            ['Total Pages', metadata.get('total_pages', 1)]
        ]
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Invoice Summary', 
                           index=False, header=False)
    
    def _create_line_items_sheet(self, data: Dict[str, Any], writer):
        """Create line items sheet in Excel."""
        line_items = data['line_items']
        
        if line_items:
            df_items = pd.DataFrame(line_items)
            df_items.to_excel(writer, sheet_name='Line Items', index=False)
        else:
            # Create empty sheet with headers
            empty_df = pd.DataFrame(columns=[
                'description', 'hsn_code', 'quantity', 'unit_price', 
                'total_amount', 'gst_rate', 'gst_amount'
            ])
            empty_df.to_excel(writer, sheet_name='Line Items', index=False)
    
    def _create_verification_sheet(self, verification_data: Dict[str, Any], writer):
        """Create verification details sheet in Excel."""
        verification_summary = [
            ['Verification Item', 'Status', 'Score/Value'],
            ['Overall Confidence', 'INFO', verification_data.get('overall_confidence', 0)],
            ['Line Item Calculations', 
             'PASS' if verification_data.get('line_item_calc_valid', True) else 'FAIL', 
             ''],
            ['Subtotal Verification', 
             'PASS' if verification_data.get('subtotal_valid', True) else 'FAIL', 
             ''],
            ['Total Verification', 
             'PASS' if verification_data.get('total_valid', True) else 'FAIL', 
             ''],
            ['Seals Detected', 
             'YES' if verification_data.get('seals_detected', False) else 'NO', 
             verification_data.get('signature_count', 0)]
        ]
        
        df_verification = pd.DataFrame(verification_summary)
        df_verification.to_excel(writer, sheet_name='Verification', 
                                index=False, header=False)
    
    def _format_excel_file(self, file_path: Path):
        """Apply formatting to Excel file."""
        try:
            wb = openpyxl.load_workbook(file_path)
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Apply header formatting if first row looks like headers
                if ws.max_row > 0:
                    first_row = ws[1]
                    for cell in first_row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", 
                                              end_color="CCCCCC", 
                                              fill_type="solid")
            
            wb.save(file_path)
            
        except Exception as e:
            logger.warning(f"Could not apply Excel formatting: {str(e)}")
    
    def _safe_float(self, value: Any) -> float:
        """Safely convert value to float."""
        try:
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                # Remove common formatting characters
                clean_value = value.replace(',', '').replace('₹', '').replace('$', '').strip()
                return float(clean_value) if clean_value else 0.0
            else:
                return 0.0
        except (ValueError, TypeError):
            return 0.0
    
    def _json_serializer(self, obj):
        """Custom JSON serializer for datetime and other objects."""
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, Path):
            return str(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


# Utility functions for external use
def create_json_output(extracted_data: Dict[str, Any], 
                      output_dir: str = "output") -> str:
    """
    Convenience function to create JSON output.
    
    Args:
        extracted_data: Extracted invoice data
        output_dir: Output directory path
        
    Returns:
        Path to created JSON file
    """
    generator = OutputGenerator(output_dir)
    return generator.create_json_output(extracted_data)


def create_excel_output(extracted_data: Dict[str, Any], 
                       output_dir: str = "output") -> str:
    """
    Convenience function to create Excel output.
    
    Args:
        extracted_data: Extracted invoice data
        output_dir: Output directory path
        
    Returns:
        Path to created Excel file
    """
    generator = OutputGenerator(output_dir)
    return generator.create_excel_output(extracted_data)


def create_verification_report(verification_data: Dict[str, Any], 
                             output_dir: str = "output") -> str:
    """
    Convenience function to create verification report.
    
    Args:
        verification_data: Verification results
        output_dir: Output directory path
        
    Returns:
        Path to created verification report
    """
    generator = OutputGenerator(output_dir)
    return generator.create_verification_report(verification_data)


# Example usage
if __name__ == "__main__":
    # Sample data for testing
    sample_extracted_data = {
        "invoice_number": "INV-2024-001",
        "invoice_date": "2024-03-15",
        "po_number": "PO-2024-100",
        "vendor_name": "ABC Technologies Pvt Ltd",
        "vendor_gst": "27ABCDE1234F1Z5",
        "buyer_name": "XYZ Corporation",
        "buyer_gst": "29XYZAB1234C1D6",
        "line_items": [
            {
                "description": "Software License",
                "hsn_code": "998361",
                "quantity": 1,
                "unit_price": 10000.00,
                "total_amount": 10000.00,
                "gst_rate": 18.0,
                "gst_amount": 1800.00
            }
        ],
        "subtotal": 10000.00,
        "discount": 0.00,
        "total_gst": 1800.00,
        "final_total": 11800.00,
        "seals_detected": True,
        "seal_count": 1
    }
    
    sample_verification_data = {
        "overall_confidence": 0.92,
        "verification_status": "VERIFIED",
        "line_item_calc_valid": True,
        "subtotal_valid": True,
        "total_valid": True,
        "seals_detected": True,
        "signature_count": 1
    }
    
    # Create output generator and generate all files
    generator = OutputGenerator()
    output_paths = generator.generate_all_outputs(
        sample_extracted_data, 
        sample_verification_data
    )
    
    print("Generated output files:")
    for file_type, path in output_paths.items():
        print(f"  {file_type}: {path}")